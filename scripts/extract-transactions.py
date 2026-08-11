# 호남THE선요양병원 1~4월 거래내역 → JSON
# 출고 12건 중 이미 등록된 4/8, 4/15, 4/16 제외하고 9건 추출.
import openpyxl, json, sys
from datetime import datetime
sys.stdout.reconfigure(encoding='utf-8')

SRC = 'C:/Users/총무구매/Documents/카카오톡 받은 파일/호남THE선요양병원26년1월~4월거래내역.xlsx'
# 이미 시스템에 등록된 입고 거래일 (엑셀 기준)
ALREADY_REGISTERED = {'2026-04-08', '2026-04-15', '2026-04-16'}

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb.active

rows = []
cur = None
for r in range(3, ws.max_row + 1):
    v = [ws.cell(row=r, column=c).value for c in range(1, 10)]
    date, kind, name, qty, price, amount, txn, paid, balance = v
    if date is not None:
        cur = { 'date': date, 'kind': kind, 'items': [] }
        rows.append(cur)
    if cur is not None and (name or qty or price or amount):
        cur['items'].append({ 'name': name, 'qty': qty, 'price': price, 'amount': amount })

# 출고만, 미등록만
out = []
for r in rows:
    if r['kind'] != '출고':
        continue
    date_str = r['date'].strftime('%Y-%m-%d')
    if date_str in ALREADY_REGISTERED:
        continue
    # 첫 행에 적힌 거래금액·결제액·잔액은 메타데이터일 뿐. items 만 추출.
    items = []
    for it in r['items']:
        # 단가/수량/금액이 모두 0/None 인 줄은 제외
        if (not it['qty'] or it['qty'] == 0) and (not it['amount'] or it['amount'] == 0):
            continue
        items.append({
            'raw_name': str(it['name'] or '').strip(),
            'qty': int(it['qty'] or 0),
            'unit_price': float(it['price'] or 0),
            'amount': float(it['amount'] or 0),
        })
    if items:
        out.append({ 'date': date_str, 'item_count': len(items), 'items': items })

print(json.dumps(out, ensure_ascii=False, indent=2))
