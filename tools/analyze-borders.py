"""셀 테두리·정렬 정밀 분석."""
import sys
import os
from openpyxl import load_workbook

def analyze(path, max_row=20):
    wb = load_workbook(path, data_only=False)
    ws = wb.worksheets[0]
    print(f'\n=== {os.path.basename(path)} | Sheet: {ws.title} ===')

    print('\n--- 셀별 정렬·테두리·폰트 (행 1-15) ---')
    for r in range(1, min(ws.max_row + 1, max_row + 1)):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if v is None:
                continue
            v_str = str(v)[:30].replace('\n', '\\n')
            font = cell.font
            align = cell.alignment
            border = cell.border

            # 테두리 정보 (style + color)
            def border_info(side):
                if side and side.style:
                    return f'{side.style}'
                return ''

            top = border_info(border.top)
            bot = border_info(border.bottom)
            left = border_info(border.left)
            right = border_info(border.right)
            border_str = ''
            if top or bot or left or right:
                border_str = f' B[T:{top or "_"}|B:{bot or "_"}|L:{left or "_"}|R:{right or "_"}]'

            align_str = ''
            if align.horizontal:
                align_str += f' h={align.horizontal}'
            if align.vertical:
                align_str += f' v={align.vertical}'
            if align.wrap_text:
                align_str += ' wrap'

            font_str = f'{font.name}/{font.size}pt'
            if font.bold:
                font_str += ' B'

            print(f'  [{cell.coordinate}] "{v_str}" | {font_str}{align_str}{border_str}')

if __name__ == '__main__':
    analyze(sys.argv[1])
