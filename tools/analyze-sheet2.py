"""두 번째 시트 (실제 결의서 페이지) 분석."""
import sys
import os
from openpyxl import load_workbook

def analyze(path):
    wb = load_workbook(path, data_only=False)
    if len(wb.worksheets) < 2:
        print('Only 1 sheet')
        return
    ws = wb.worksheets[1]
    print(f'\n=== {os.path.basename(path)} | Sheet[1]: {ws.title} ===')
    print(f'Dim: {ws.max_row} rows × {ws.max_column} cols')

    # 컬럼 너비
    widths = []
    for col_letter in 'ABCDEFGHIJKLMN'[:ws.max_column]:
        cd = ws.column_dimensions.get(col_letter)
        if cd and cd.width:
            widths.append(f'{col_letter}={cd.width:.1f}')
    print(f'Column widths: {", ".join(widths)}')

    # 행 높이
    heights = []
    for r in range(1, min(ws.max_row + 1, 80)):
        rd = ws.row_dimensions.get(r)
        if rd and rd.height:
            heights.append((r, rd.height))
    print(f'Row heights: {heights[:15]}')

    # 페이지 설정
    pm = ws.page_margins
    print(f'Margins(in): t={pm.top:.3f} b={pm.bottom:.3f} l={pm.left:.3f} r={pm.right:.3f}')

    # 병합 셀
    merged = sorted([str(m) for m in ws.merged_cells.ranges])
    print(f'\nMerged cells ({len(merged)}):')
    for m in merged[:50]:
        print(f'  {m}')

    # 헤더 1-15행
    print('\n--- Header rows (1-15) with full style ---')
    for r in range(1, 16):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if v is None:
                continue
            v_str = str(v)[:60].replace('\n', '\\n')
            font = cell.font
            align = cell.alignment
            border = cell.border
            fill = cell.fill

            def bs(s):
                return s.style if s and s.style else '_'

            border_str = f'B[T:{bs(border.top)}|B:{bs(border.bottom)}|L:{bs(border.left)}|R:{bs(border.right)}]'

            font_str = f'{font.name}/{font.size}pt'
            if font.bold:
                font_str += ' B'

            align_str = ''
            if align.horizontal:
                align_str += f' h={align.horizontal}'
            if align.vertical:
                align_str += f' v={align.vertical}'

            fill_str = ''
            if fill.fgColor and fill.fgColor.rgb and fill.fgColor.rgb != '00000000':
                fill_str = f' fill={fill.fgColor.rgb}'

            print(f'  [{cell.coordinate}] "{v_str}" | {font_str}{align_str}{border_str}{fill_str}')

    # 표 영역 - 데이터 행 1-2개 + 합계
    print('\n--- Sample data rows (15-25) ---')
    for r in range(15, min(ws.max_row + 1, 26)):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if v is None:
                continue
            v_str = str(v)[:30].replace('\n', '\\n')
            print(f'  [{cell.coordinate}] "{v_str}"')

    # 마지막 영역 (합계)
    print(f'\n--- Last 10 rows ({ws.max_row - 10}-{ws.max_row}) ---')
    for r in range(max(1, ws.max_row - 10), ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if v is None:
                continue
            v_str = str(v)[:50].replace('\n', '\\n')
            font = cell.font
            font_str = f'{font.name}/{font.size}pt'
            if font.bold:
                font_str += ' B'
            print(f'  [{cell.coordinate}] "{v_str}" {font_str}')

if __name__ == '__main__':
    analyze(sys.argv[1])
