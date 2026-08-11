"""구매결의서 양식 파일 분석.
xls/xlsx 파일에서 셀 위치, 너비, 높이, 폰트, 정렬, 병합, 테두리, 색상 등 추출.
"""
import sys
import os

def analyze_xls(path):
    """xls (older) 파일 분석"""
    import xlrd
    book = xlrd.open_workbook(path, formatting_info=True)
    print(f'\n=== {os.path.basename(path)} (xls) ===')
    for sheet in book.sheets():
        print(f'\n--- Sheet: {sheet.name} ({sheet.nrows} rows × {sheet.ncols} cols) ---')
        # 컬럼 너비 (일반 단위 1/256 of average char width)
        widths = []
        for c in range(sheet.ncols):
            try:
                w = sheet.computed_column_width(c)
                widths.append(round(w, 1))
            except Exception:
                widths.append(None)
        print(f'Column widths: {widths}')
        # 병합 셀
        merged = sheet.merged_cells
        if merged:
            print(f'Merged cells (row1,row2,col1,col2): {merged}')
        # 셀 내용
        for r in range(min(sheet.nrows, 50)):
            row_data = []
            for c in range(sheet.ncols):
                v = sheet.cell_value(r, c)
                xf_idx = sheet.cell_xf_index(r, c)
                xf = book.xf_list[xf_idx]
                font = book.font_list[xf.font_index]
                align = xf.alignment
                if isinstance(v, str) and v.strip():
                    row_data.append(f'[{r},{c}] "{v[:50]}" font={font.name}/{font.height/20}pt bold={font.bold} align_h={align.hor_align} align_v={align.vert_align}')
                elif v != '' and v != 0:
                    row_data.append(f'[{r},{c}] {v}')
            if row_data:
                for d in row_data:
                    print('  ', d)

def analyze_xlsx(path):
    """xlsx (newer) 파일 분석"""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=False)
    print(f'\n=== {os.path.basename(path)} (xlsx) ===')
    for ws in wb.worksheets:
        print(f'\n--- Sheet: {ws.title} ({ws.max_row} rows × {ws.max_column} cols) ---')
        # 컬럼 너비
        widths = []
        for col_letter in 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'[:ws.max_column]:
            cd = ws.column_dimensions.get(col_letter)
            if cd and cd.width:
                widths.append((col_letter, round(cd.width, 1)))
        print(f'Column widths: {widths}')
        # 행 높이 (일반 행 제외, 다른 것만)
        heights = []
        for r in range(1, min(ws.max_row + 1, 50)):
            rd = ws.row_dimensions.get(r)
            if rd and rd.height:
                heights.append((r, rd.height))
        print(f'Row heights (custom only): {heights[:10]}')
        # 병합 셀
        merged = list(ws.merged_cells.ranges)
        if merged:
            print(f'Merged cells: {[str(m) for m in merged[:20]]}')
        # 페이지 설정
        ps = ws.page_setup
        pm = ws.page_margins
        print(f'Page: {ps.paperSize} orientation={ps.orientation}')
        print(f'Margins: top={pm.top} bottom={pm.bottom} left={pm.left} right={pm.right}')
        # 셀 내용 + 스타일
        max_r = min(ws.max_row, 60)
        for r in range(1, max_r + 1):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                v = cell.value
                if v is None or (isinstance(v, str) and not v.strip()):
                    continue
                font = cell.font
                align = cell.alignment
                fill = cell.fill
                v_str = str(v)[:60].replace('\n', '\\n')
                style_parts = []
                if font.name:
                    style_parts.append(f'font={font.name}/{font.size}pt')
                if font.bold:
                    style_parts.append('bold')
                if align.horizontal:
                    style_parts.append(f'h={align.horizontal}')
                if align.vertical:
                    style_parts.append(f'v={align.vertical}')
                if fill.fgColor and fill.fgColor.rgb and fill.fgColor.rgb != '00000000':
                    style_parts.append(f'fill={fill.fgColor.rgb}')
                style = ' '.join(style_parts)
                print(f'  [{cell.coordinate}] "{v_str}" {style}')

if __name__ == '__main__':
    target = sys.argv[1] if len(sys.argv) > 1 else None
    if not target:
        print('Usage: python analyze-xls.py <file>')
        sys.exit(1)
    if target.lower().endswith('.xlsx'):
        analyze_xlsx(target)
    else:
        analyze_xls(target)
