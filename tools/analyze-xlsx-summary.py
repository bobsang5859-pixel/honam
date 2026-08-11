"""구매결의서 양식 요약 — 첫 시트만, 핵심 구조."""
import sys
import os
from openpyxl import load_workbook

def analyze(path):
    wb = load_workbook(path, data_only=False)
    print(f'\n=== {os.path.basename(path)} ===')
    print(f'Sheets: {wb.sheetnames[:3]}{"..." if len(wb.sheetnames) > 3 else ""}  (total {len(wb.sheetnames)})')

    # 첫 시트만 분석
    ws = wb.worksheets[0]
    print(f'\n--- Sheet[0]: {ws.title} ---')
    print(f'Dim: {ws.max_row} rows × {ws.max_column} cols')

    # 컬럼 너비
    widths = []
    for col_letter in 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'[:ws.max_column]:
        cd = ws.column_dimensions.get(col_letter)
        if cd and cd.width:
            widths.append(f'{col_letter}={cd.width:.1f}')
    print(f'Column widths: {", ".join(widths)}')

    # 페이지 설정
    ps = ws.page_setup
    pm = ws.page_margins
    print(f'Paper: size={ps.paperSize} orient={ps.orientation} fitTo={ps.fitToWidth}x{ps.fitToHeight}')
    print(f'Margins(in): t={pm.top} b={pm.bottom} l={pm.left} r={pm.right}')

    # 병합 셀
    merged = sorted([str(m) for m in ws.merged_cells.ranges])
    print(f'Merged cells ({len(merged)}): {merged[:30]}')

    # 헤더 영역 (1~12행 모든 셀)
    print('\n--- Header rows (1-12) ---')
    for r in range(1, 13):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if v is None:
                continue
            v_str = str(v)[:60].replace('\n', '\\n')
            font = cell.font
            align = cell.alignment
            style = f'{font.name}/{font.size}pt'
            if font.bold: style += ' bold'
            if align.horizontal: style += f' h={align.horizontal}'
            print(f'  [{cell.coordinate}] "{v_str}" {style}')

    # 데이터 행 첫 5개 (스타일 확인)
    print('\n--- Data rows (first 3 with content after row 12) ---')
    count = 0
    for r in range(13, ws.max_row + 1):
        if count >= 3: break
        has_content = False
        row_cells = []
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if v is None: continue
            v_str = str(v)[:30].replace('\n', '\\n')
            row_cells.append(f'{cell.coordinate}="{v_str}"')
            has_content = True
        if has_content:
            count += 1
            print(f'  R{r}: {", ".join(row_cells)}')

    # Footer/총합계 위치
    print('\n--- Total/footer rows (search "합계", "총") ---')
    for r in range(max(1, ws.max_row - 20), ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if v and isinstance(v, str) and ('합계' in v or '총' in v or '계' in v or '소 계' in v):
                font = cell.font
                style = f'{font.name}/{font.size}pt'
                if font.bold: style += ' bold'
                print(f'  [{cell.coordinate}] "{v}" {style}')

if __name__ == '__main__':
    analyze(sys.argv[1])
